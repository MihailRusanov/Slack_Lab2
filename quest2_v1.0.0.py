from openpyxl import load_workbook

wb = load_workbook('./students.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
quantity_of_students = sheet.max_row - 1


# Функция сортировки списка студентов
def get_sorted_list(sheet):
    mas = []
    element = 'A'
    j = 2
    for i in range(quantity_of_students):
        element = 'A' + str(j)
        mas.append(list(map(str, sheet[element].value.split())))
        j += 1

    element = 'B'
    i = 0
    for i in range(quantity_of_students):
        element = 'B' + str(i + 2)
        mas[i].append(sheet[element].value)
        i += 1
    res = sorted(mas, key=lambda x: x[3])

    return res

# вывод таблицы
def print_table(res):
    print("ФИО студента - его дата рождения")
    i = 0
    while i < quantity_of_students:
        print(res[i][0], res[i][1], res[i][2], "-", res[i][3], sep=' ')
        i += 1

print_table(get_sorted_list(sheet))
