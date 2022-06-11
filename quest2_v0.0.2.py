from openpyxl import load_workbook

wb = load_workbook('./students.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
quantity_of_students = sheet.max_row - 1

mas = []
element = 'A'
j = 2
for i in range(quantity_of_students):
    element = 'A' + str(j)
    mas.append(list(map(str, sheet[element].value.split())))
    j += 1

print(mas)
