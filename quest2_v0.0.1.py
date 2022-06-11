from openpyxl import load_workbook

wb = load_workbook('./students.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
quantity_of_students = sheet.max_row - 1

print(quantity_of_students)
