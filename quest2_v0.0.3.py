from openpyxl import load_workbook

wb = load_workbook('./students.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
quantity_of_students = sheet.max_row - 1

def get_sorted_list(sheet):
    mas = []
    element = 'A'
    j = 2
    for i in range(quantity_of_students):
        element = 'A' + str(j)
        mas.append(list(map(str, sheet[element].value.split())))
        j += 1

    element = 'B'
    i = 0
    for i in range(quantity_of_students):
        element = 'B' + str(i + 2)
        mas[i].append(sheet[element].value)
        i += 1
    res = sorted(mas, key=lambda x: x[3])

    return res

print(get_sorted_list(sheet))
